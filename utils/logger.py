import logging
import os
import csv
import pandas as pd
from datetime import datetime
from config.config import PathConfig
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows



def setup_logger(name: str, log_file: str, level=logging.INFO) -> logging.Logger:
    """
    로거를 설정하고 반환합니다.
    """
    logger = logging.getLogger(name)
    
    if logger.handlers:
        return logger
    
    logger.setLevel(level)

    # Log 폴더 없으면 생성
    os.makedirs(os.path.dirname(log_file), exist_ok=True)

    handler = logging.FileHandler(log_file, encoding='utf-8')
    formatter = logging.Formatter('%(asctime)s | %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
    handler.setFormatter(formatter)

    logger.addHandler(handler)

    return logger


def write_log(message: str, filename: str):
    """
    로그 메시지를 텍스트 파일에 기록합니다.
    :param message: 로그 메시지
    :param filename: 로그 파일 이름
    """
    os.makedirs(PathConfig.RESULT_DIR, exist_ok=True)

    # ✅ log_dir과 filename을 합쳐서 경로 구성
    file_path = os.path.join(PathConfig.RESULT_DIR, filename)

    with open(file_path, "a", encoding="utf-8") as f:
        f.write(message + "\n")


def write_log_csv(row:dict, filename: str):
    """
    로그 메시지를 CSV 파일에 기록합니다.
    :param row: 로그 메시지 (딕셔너리 형태)
    :param filename: 로그 파일 이름
    """
    os.makedirs(PathConfig.RESULT_DIR, exist_ok=True)

    # ✅ log_dir과 filename을 합쳐서 경로 구성
    file_path = os.path.join(PathConfig.RESULT_DIR, filename)

    # CSV 파일에 헤더가 없으면 추가
    file_exists = os.path.isfile(file_path)

    with open(file_path, mode='a', newline='', encoding='utf-8-sig') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=row.keys())
        if not file_exists:
            writer.writeheader()
        writer.writerow(row)


def write_log_xlsx(df: pd.DataFrame, filename: str, template: str = "default"):
    """
    DataFrame 전체를 XLSX 파일에 저장합니다. 기존 파일은 덮어씌워집니다.
    :param df: 기록할 DataFrame
    :param filename: 로그 파일 이름 (xlsx 확장자 포함)
    :param template: "score", "trading" 등 포맷 유형에 따른 열 너비 지정
    """
    os.makedirs(PathConfig.RESULT_DIR, exist_ok=True)
    file_path = os.path.join(PathConfig.RESULT_DIR, filename)

    wb = Workbook()
    ws = wb.active

    # 🧾 헤더 작성
    header = list(df.columns)
    ws.append(header)

    for col_num, col_name in enumerate(header, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = Font(name="NanumGothic", size=9, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

    # 🔧 템플릿별 열 너비 설정
    if template == "score":
        col_widths = [14, 12, 10, 10, 10, 12]
    elif template == "trading":
        col_widths = [14, 14, 10, 12, 10]
    else:
        col_widths = [15] * len(header)

    for i, width in enumerate(col_widths, 1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = width

    # 🔍 필터 추가 및 틀 고정
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    # ✏️ 본문 데이터 작성
    for row_data in df.itertuples(index=False, name=None):
        ws.append(row_data)
        for i, value in enumerate(row_data):
            cell = ws.cell(row=ws.max_row, column=i + 1)
            cell.font = Font(name="NanumGothic", size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # 💾 저장
    wb.save(file_path)
